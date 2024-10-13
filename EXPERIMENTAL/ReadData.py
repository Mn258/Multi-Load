import openpyxl
import pickle

# XLSX文件路径
xlsx_file = 'EXPERIMENTAL/repositioning.xlsx'
# xlsx_file = 'result_data.xlsx'

# 加载工作簿
wb = openpyxl.load_workbook(xlsx_file)

# 获取所有工作表的名称
sheet_names = wb.sheetnames

total_group = {}
# 遍历所有工作表
for sheet_name in sheet_names[:7]:
    # 根据工作表名称获取工作表对象
    sheet = wb[sheet_name]

    # 打印工作表名称
    print(f'工作表名称: {sheet_name}' + 's')

    temp_one_group = {"metric": [], "coordinate": []}
    # 遍历并打印工作表的每一行数据
    coordinate = []
    for row in sheet.iter_rows(values_only=True):
        print(row)
        if row[0] is None:
            if sheet_name == "taskNum":
                for i in range(len(row)):
                    if i==0:
                        continue
                    temp_one_group["coordinate"].append(str(row[i]//1000)+'k')
                    coordinate.append(str(int(row[i]/1000))+'k')
            else:
                temp_one_group["coordinate"] = row[1:]
                coordinate = row[1:]
        elif row[1] is None:
            temp_one_group["metric"] = row[0]
        else:
            if temp_one_group["metric"] == "Makespan":
                temp_sum_of_cost = []
                print(row[1:], "??????>>>>>>>")
                for i in row[1:]:
                    # pri
                    if i is not None:
                        temp_sum_of_cost.append(i/1000)
                temp_one_group[row[0]] = temp_sum_of_cost
            elif temp_one_group["metric"] == "Service Time":
                temp_sum_of_cost = []
                print(row[1:], "??????>>>>>>>")
                for i in row[1:]:
                    # pri
                    if i is not None:
                        temp_sum_of_cost.append(i/10000)
                temp_one_group[row[0]] = temp_sum_of_cost
            else:
                temp_one_group[row[0]] = row[1:]
        if len(temp_one_group) == 6 : # or ((sheet_name == 'T' or temp_one_group["metric"] == "Processing Time")  and len(temp_one_group) == 6)
            print(temp_one_group, "?????????>!!!")
            print(temp_one_group["metric"], "?????>>>>>>>>>")
            total_group[sheet_name+temp_one_group["metric"]] = temp_one_group
            temp_one_group = {"metric": [], "coordinate": coordinate}
        # assert 1 == 2
    print()  # 打印空行，用于分隔不同的工作表内容

print("=====================")
print(total_group)
print("---------------------")
# 关闭工作簿
wb.close()

with open(xlsx_file +'.pkl', 'wb') as file:
    pickle.dump(total_group, file)